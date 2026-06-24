# services/core/src/routers/cash.py
from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from pydantic import BaseModel
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from src.database import get_db
from src.models import (
    CashDay,
    CashEvent,
    CashEventItem,
    CashEventType,
    ProductUnit,
    Product,
)
from src.components.cash_day_manager import CashDayManager
from src.components.sales_manager import SalesManager
from src.components.return_manager import ReturnManager
from src.components.cash_adjustments import get_sold_units, adjust_sale

router = APIRouter(prefix="/cash", tags=["Кассовый узел и Продажи"])


@router.get("/days/current/export-excel")
async def export_current_shift_to_excel(db: AsyncSession = Depends(get_db)):
    """📊 Сбор данных текущей смены и генерация отчета Excel"""
    day_res = await db.execute(select(CashDay).where(CashDay.is_closed == False))
    active_day = day_res.scalar_one_or_none()
    if not active_day:
        raise HTTPException(
            status_code=400, detail="Нет открытой смены для выгрузки отчета"
        )

    stmt = (
        select(CashEvent, CashEventItem, ProductUnit, Product)
        .join(CashEventItem, CashEventItem.cash_event_id == CashEvent.id)
        .join(ProductUnit, ProductUnit.id == CashEventItem.product_unit_id)
        .join(Product, Product.id == ProductUnit.product_id)
        .where(
            CashEvent.cash_day_id == active_day.id,
            CashEvent.type.in_([CashEventType.SALE, CashEventType.RETURN]),
        )
        .order_by(CashEvent.created_at.asc())
    )

    query_res = await db.execute(stmt)
    records = query_res.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Кассовая смена"
    ws.sheet_properties.showGridLines = True

    font_title = Font(name="Arial", size=14, bold=True, color="333333")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=11, bold=False)
    font_total = Font(name="Arial", size=11, bold=True)

    fill_header = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    fill_return = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    shift_date = active_day.created_at.strftime("%d.%m.%Y")
    ws.append(
        [f"Журнал продаж за кассовую смену от {shift_date} (Смена #{active_day.id})"]
    )
    ws.cell(row=1, column=1).font = font_title
    ws.append([])

    headers = [
        "Событие",
        "ID",
        "Артикул",
        "Наименование",
        "Остаток",
        "Стоимость",
        "Примечание",
        "Log",
        "Заказная",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_row = 4
    total_sum = 0.0

    for row_data in records:
        event, item, unit, product = row_data

        event_type = "Продажа" if event.type == CashEventType.SALE else "Возврат"
        amount = float(event.total_amount)

        if event.type == CashEventType.RETURN:
            total_sum -= amount
        else:
            total_sum += amount

        # Количество IN_STORE для этого товара
        stock_qty = (
            await db.execute(
                select(func.count(ProductUnit.id)).where(
                    ProductUnit.product_id == product.id,
                    ProductUnit.physical_status == "IN_STORE",
                )
            )
        ).scalar_one() or 0

        row_values = [
            event_type,
            f"#{event.id}",
            product.code if product else "—",
            product.name.replace("_", " ") if product else "—",
            stock_qty,
            amount,
            event.description or "",
            "",  # Log — пока пусто
            "",  # Заказная — пока пусто
        ]

        ws.append(row_values)

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if event.type == CashEventType.RETURN:
                cell.fill = fill_return
            if col_idx in (1, 2, 5):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

        current_row += 1

    ws.append([])
    current_row += 1
    total_row = ["", "", "", "", "ИТОГО ПО СМЕНЕ:", total_sum, "", "", ""]
    ws.append(total_row)
    for col_idx in range(1, len(total_row) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = font_total
        if col_idx == 6:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    for i, col in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = chr(64 + i) if i <= 26 else "A"
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"report_shift_{active_day.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/days/current/sales", status_code=200)
async def current_sales(db: AsyncSession = Depends(get_db)):
    day_stmt = select(CashDay).where(CashDay.is_closed == False)
    day_res = await db.execute(day_stmt)
    active_day = day_res.scalar_one_or_none()
    if not active_day:
        return {"sales": [], "message": "Нет открытой смены"}

    events = (
        (
            await db.execute(
                select(CashEvent)
                .where(
                    CashEvent.cash_day_id == active_day.id,
                    CashEvent.type.in_([CashEventType.SALE, CashEventType.RETURN]),
                )
                .order_by(CashEvent.created_at.asc())
            )
        )
        .scalars()
        .all()
    )

    sales = []
    for event in events:
        items = (
            (
                await db.execute(
                    select(CashEventItem).where(CashEventItem.cash_event_id == event.id)
                )
            )
            .scalars()
            .all()
        )
        for item in items:
            unit = await db.get(ProductUnit, item.product_unit_id)
            product = await db.get(Product, unit.product_id) if unit else None
            sales.append(
                {
                    "time": event.created_at.isoformat(),
                    "product_name": product.name if product else "Товар",
                    "price": float(event.total_amount),
                    "serial_number": unit.unique_serial_number if unit else "—",
                    "event_id": event.id,
                    "type": event.type.value,
                    "description": event.description or "",
                }
            )
    return {"sales": sales, "total": len(sales)}


@router.get("/days/current/units", status_code=200)
async def sold_units(
    days: int = Query(5, ge=1, le=365), db: AsyncSession = Depends(get_db)
):
    return await get_sold_units(days, db)


@router.get("/days", status_code=200)
async def list_cash_days(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(CashDay).order_by(CashDay.created_at.desc()).limit(100)
    )
    days = result.scalars().all()
    return {
        "days": [
            {
                "id": d.id,
                "created_at": str(d.created_at),
                "status": "ЗАКРЫТА" if d.is_closed else "ОТКРЫТА",
                "total_sales": float(d.total_revenue),
            }
            for d in days
        ],
        "total": len(days),
    }


class CashReturnPayload(BaseModel):
    unique_serial_number: str
    return_reason: str = "Возврат от покупателя"


@router.post("/days/open", status_code=201)
async def open_day(payload: dict = Body(...), db: AsyncSession = Depends(get_db)):
    return await CashDayManager.open_day(payload, db)


@router.post("/sales", status_code=201)
async def process_sale(payload: dict = Body(...), db: AsyncSession = Depends(get_db)):
    product_id = payload.get("product_id")
    if not product_id:
        raise HTTPException(status_code=422, detail="product_id обязателен")
    return await SalesManager.execute_fifo_sale(int(product_id), payload, db)


@router.post("/sales/adjust", status_code=200)
async def adjust(payload: dict, db: AsyncSession = Depends(get_db)):
    return await adjust_sale(payload, db)


@router.post("/days/{cash_day_id}/reopen", status_code=200)
async def reopen_day(cash_day_id: int, db: AsyncSession = Depends(get_db)):
    return await CashDayManager.reopen_day(cash_day_id, db)


@router.post("/days/close", status_code=200)
async def close_day(db: AsyncSession = Depends(get_db)):
    return await CashDayManager.close_day(db)


@router.get("/returns/check-relation", status_code=200)
async def check_return(sn: str = Query(...), db: AsyncSession = Depends(get_db)):
    return await ReturnManager.check_unit_relation(sn, db)


@router.post("/returns", status_code=200)
async def process_return(
    payload: CashReturnPayload, db: AsyncSession = Depends(get_db)
):
    day = (
        await db.execute(select(CashDay).where(CashDay.is_closed == False))
    ).scalar_one_or_none()
    if not day:
        raise HTTPException(status_code=400, detail="Смена не открыта")
    return await ReturnManager.execute_return(
        payload.unique_serial_number, payload.return_reason, day.id, db
    )
